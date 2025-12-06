import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

from docx import Document
from openpyxl import Workbook, load_workbook

from resume_customizer import customize_resume_with_placeholders

import subprocess

excel_log_path = "/Users/Roger/Documents/FullTime-Resume/Job Tracker.xlsx"


def convert_docx_to_pdf(docx_path: str) -> str:
    """
    Convert a DOCX file to PDF in the same folder.
    Tries methods in order: WPS Office (via AppleScript), LibreOffice.
    Returns the path to the PDF file if successful, None otherwise.
    This function will NEVER raise exceptions - all errors are caught and logged.
    """
    try:
        docx_path_obj = Path(docx_path)
        if not docx_path_obj.exists():
            print(f"⚠️ DOCX file not found: {docx_path}")
            return None

        # Output PDF path (same folder, same name, .pdf extension)
        pdf_path = docx_path_obj.with_suffix('.pdf')
        output_dir = docx_path_obj.parent
        docx_abs_path = str(docx_path_obj.resolve())

        # Method 1: Try WPS Office via AppleScript
        wps_app_paths = [
            "/Applications/WPS Office.app",
            "/Applications/Kingsoft Office.app",
        ]

        wps_found = False
        for wps_app in wps_app_paths:
            if Path(wps_app).exists():
                wps_found = True
                print(f"📝 Attempting PDF conversion with WPS Office...")

                try:
                    applescript = f'''
                    tell application "WPS Office"
                        activate
                        open POSIX file "{docx_abs_path}"
                    end tell
                    delay 4
                    tell application "System Events"
                        tell process "WPS Office"
                            try
                                keystroke "s" using {{command down, shift down}}
                                delay 2
                                keystroke return
                                delay 2
                            on error errMsg
                                return "Error: " & errMsg
                            end try
                        end tell
                    end tell
                    delay 3
                    tell application "WPS Office"
                        quit
                    end tell
                    '''

                    result = subprocess.run(
                        ["osascript", "-e", applescript],
                        capture_output=True,
                        text=True,
                        timeout=30
                    )

                    # If expected PDF exists
                    if pdf_path.exists():
                        print(f"✅ PDF exported using WPS Office: {pdf_path}")
                        return str(pdf_path)

                    # Otherwise, try to find most recent PDF and rename
                    pdf_files = list(output_dir.glob("*.pdf"))
                    if pdf_files:
                        recent_pdf = max(pdf_files, key=lambda p: p.stat().st_mtime)
                        print(f"⚠️ WPS Office created PDF but with different name: {recent_pdf}")
                        recent_pdf.rename(pdf_path)
                        print(f"✅ Renamed to: {pdf_path}")
                        return str(pdf_path)

                    print(f"⚠️ WPS Office automation attempted but PDF not created")
                    if result.stderr:
                        print(f"   Error: {result.stderr.strip()[:200]}")
                except (subprocess.TimeoutExpired, FileNotFoundError, Exception) as e:
                    print(f"⚠️ WPS Office conversion failed: {e}")
                break

        # Method 2: Try LibreOffice (command line)
        if not pdf_path.exists():
            libreoffice_paths = [
                "/Applications/LibreOffice.app/Contents/MacOS/soffice",
                "/usr/local/bin/soffice",
                "/opt/homebrew/bin/soffice",
                shutil.which("soffice"),
            ]

            libreoffice_cmd = None
            for path in libreoffice_paths:
                if path and Path(path).exists():
                    libreoffice_cmd = path
                    break

            if libreoffice_cmd:
                print(f"📝 Attempting PDF conversion with LibreOffice...")
                try:
                    cmd = [
                        libreoffice_cmd,
                        "--headless",
                        "--convert-to", "pdf",
                        "--outdir", str(output_dir),
                        str(docx_path_obj)
                    ]
                    result = subprocess.run(
                        cmd,
                        capture_output=True,
                        text=True,
                        timeout=60
                    )

                    if result.returncode == 0 and pdf_path.exists():
                        print(f"✅ PDF exported using LibreOffice: {pdf_path}")
                        return str(pdf_path)
                    elif result.stderr:
                        print(f"⚠️ LibreOffice conversion failed: {result.stderr[:200]}")
                except (subprocess.TimeoutExpired, FileNotFoundError, Exception) as e:
                    print(f"⚠️ LibreOffice conversion failed: {e}")

        # All methods failed - non-critical
        if not pdf_path.exists():
            if wps_found:
                print(f"⚠️ PDF conversion unsuccessful. WPS automation needs improvement.")
                print(f"💡 DOCX is available for manual PDF export.")
            else:
                print(f"⚠️ PDF conversion skipped: No compatible office suite found.")
                print(f"💡 Tip: Install LibreOffice: brew install --cask libreoffice")

        return None

    except Exception as e:
        print(f"⚠️ Unexpected error in PDF conversion (non-critical): {e}")
        print(f"   DOCX file is still available: {docx_path}")
        return None


def log_application_to_excel(excel_path, sheet_name, company_name, position_name, applied_date, job_description=None):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(["Company", "Position", "Applied Date", "Job Description"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Company", "Position", "Applied Date", "Job Description"])

    ws.append([company_name, position_name, applied_date, job_description or ""])
    print(f"➡️ Logging row: {company_name}, {position_name}, {applied_date}")
    wb.save(excel_path)
    print(f"💾 Saved to {excel_path}")


def replace_placeholders_in_docx(input_path, output_path, replacements: dict):
    """Replace placeholders like {{COMPANY_NAME}} in a Word document."""
    doc = Document(input_path)
    for para in doc.paragraphs:
        for key, value in replacements.items():
            if key in para.text:
                for run in para.runs:
                    if key in run.text:
                        run.text = run.text.replace(key, value)
    doc.save(output_path)


def create_application_folder(
    company_name,
    position_name,
    resume_path,
    coverLetter_path,
    location,
    jd_source_path=None,
    job_category="sde",
    position_type=None,
):
    """
    Create an application folder for a company/position, copy resume & cover letter,
    copy job description file, and optionally customize resume with LLM.

    Args:
        job_category: "sde" or "support" - determines sections to use
        position_type: "fullstack" or "support" – overrides job_category if provided
    """
    # --- DEBUG ---
    print(f"[DEBUG create_application_folder] job_category={job_category!r}, position_type(before)={position_type!r}")

    # Derive position_type from job_category if not explicitly given
    if position_type is None:
        position_type = "fullstack" if job_category == "sde" else "support"

    print(f"[DEBUG create_application_folder] position_type(final)={position_type!r}")

    parent_folder = os.path.dirname(resume_path)
    grandparent_folder = os.path.dirname(parent_folder)

    # Create folders
    company_folder = os.path.join(grandparent_folder, company_name)
    os.makedirs(company_folder, exist_ok=True)
    position_folder = os.path.join(company_folder, position_name)
    os.makedirs(position_folder, exist_ok=True)

    # Define filenames
    resume_filename = f"Roger Xu_{company_name}_Resume_Template.docx"
    cover_filename = f"Roger Xu_{company_name}_CoverLetter_Template.docx"
    jd_filename = "job_description.txt"

    resume_target = os.path.join(position_folder, resume_filename)
    cover_target = os.path.join(position_folder, cover_filename)
    jd_target = os.path.join(position_folder, jd_filename)

    # Date string
    today_str = datetime.today().strftime("%B %d, %Y")

    # Placeholder replacements
    replacements = {
        "{{COMPANY_NAME}}": company_name,
        "{{POSITION_NAME}}": position_name,
        "{{TODAY_DATE}}": today_str,
        "LOCATION_STRING": location
    }

    # Step 1: Copy resume template
    shutil.copy(resume_path, resume_target)

    # Step 2: Copy JD and read additional_info if present
    job_description = None
    additional_info = None

    if jd_source_path and os.path.exists(jd_source_path):
        shutil.copy(jd_source_path, jd_target)
        with open(jd_target, "r") as f:
            job_description = f.read()

        addl_source_path = os.path.join(os.path.dirname(jd_source_path), "additional_info.txt")
        if os.path.exists(addl_source_path):
            with open(addl_source_path, "r") as f:
                additional_info = f.read()
            print("ℹ️ Loaded additional_info.txt")

    # Step 3: Customize resume
    position_type_folder_name = position_type.capitalize() + "_Sections"
    print(f"[DEBUG] Section folder: {position_type_folder_name}")

    # Build section_files
    section_files = {
        "SUMMARY": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/summary.txt",
        "HOOPP_EXPERIENCE": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/hoopp_experience.txt",
        "SKILLS": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/skills.txt",
    }

    # Add JOBPILOT and PORTFOLIO_TRACKER only for SDE jobs
    if job_category == "sde":
        section_files["PORTFOLIO_TRACKER"] = f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/portfolio_tracker.txt"
        section_files["JOBPILOT"] = f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/jobpilot.txt"

    if job_description:
        customized_resume_path = os.path.join(
            position_folder,
            f"Roger Xu_{company_name}_Resume.docx"
        )
        print("📄 Customizing resume with LLM...")
        customize_resume_with_placeholders(
            resume_path,
            section_files,
            job_description,
            customized_resume_path,
            job_category,
            additional_info
        )
        resume_target = customized_resume_path

        # Export resume as PDF
        print("📄 Exporting resume to PDF...")
        try:
            convert_docx_to_pdf(customized_resume_path)
        except Exception as e:
            print(f"⚠️ PDF export failed (non-critical): {e}")
            print("   Resume DOCX file is still available for manual PDF export.")

    # Step 4: Generate cover letter (template with placeholders)
    replace_placeholders_in_docx(coverLetter_path, cover_target, replacements)

    print(f"✅ Application folder created: {position_folder}")
    print(f"📄 Files created so far:\n- {resume_target}\n- {cover_target}\n- {jd_target if job_description else 'No JD copied (JD file missing)'}")

    # Step 4.5: Log to Excel
    log_application_to_excel(
        excel_log_path,
        sheet_name="Job Tracker",
        company_name=company_name,
        position_name=position_name,
        applied_date=today_str,
        job_description=job_description,
    )
    print(f"📝 Logged application to {excel_log_path}")

    # Step 5: Clear only the source job_description.txt and additional_info.txt
    try:
        if jd_source_path and os.path.exists(jd_source_path):
            open(jd_source_path, "w").close()
            print("🧹 Cleared source job_description.txt content")

        addl_source_path = os.path.join(os.path.dirname(jd_source_path), "additional_info.txt") if jd_source_path else None
        if addl_source_path and os.path.exists(addl_source_path):
            open(addl_source_path, "w").close()
            print("🧹 Cleared source additional_info.txt content")
    except Exception as e:
        print(f"⚠️ Failed to clear source job description/additional_info: {e}")

    # Step 6: Record cover letter absolute path
    try:
        record_cover_path = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/cover_letter_path.txt"
        with open(record_cover_path, "w") as f:
            f.write(cover_target)
        print(f"📝 Recorded cover letter path: {cover_target} → {record_cover_path}")
    except Exception as e:
        print(f"⚠️ Failed to write cover letter path: {e}")

    return True


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python script.py <CompanyName> <PositionName> [easy_apply] [location] [job_category]")
        sys.exit(1)

    # Resume templates
    resume_default = "/Users/Roger/Documents/FullTime-Resume/Rong Gang Xu_Resume_v3.docx"
    resume_frontend = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/Roger Xu_Frontend_Resume_Placeholder.docx"
    resume_fullstack = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/Roger Xu_Fullstack_Resume_Placeholder.docx"
    resume_support = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/Roger Xu_Support_Resume_Placeholder.docx"  # Using support-specific template
    resume_sharepoint = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/Roger Xu_SharePoint_Resume.docx"
    coverLetter = "/Users/Roger/Documents/FullTime-Resume/Roger Xu_coverletter.docx"

    jd_source_path = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/job_description.txt"

    # CLI args
    company = sys.argv[1].strip()
    position = sys.argv[2].strip()

    # Optional args with safe defaults
    easy_apply = sys.argv[3].strip().lower() if len(sys.argv) > 3 else "true"
    location = sys.argv[4].strip() if len(sys.argv) > 4 else "Toronto, ON"
    job_category = sys.argv[5].strip().lower() if len(sys.argv) > 5 else "sde"

    print("======== DEBUG ARGS ========")
    print("sys.argv:", sys.argv)
    print("company:", company)
    print("position:", position)
    print("easy_apply:", easy_apply)
    print("location:", location)
    print("job_category:", job_category)
    print("============================")

    # Derive position_type from job_category
    position_type = "fullstack" if job_category == "sde" else "support"
    print(f"[DEBUG main] position_type={position_type!r}")

    # Select resume template based on position_type
    if position_type == "fullstack":
        resume = resume_fullstack
    elif position_type == "support":
        resume = resume_support
    else:
        resume = resume_fullstack  # fallback

    print(f"[DEBUG main] Using resume template: {resume!r}")

    # Create folder & base files
    folder_created = create_application_folder(
        company,
        position,
        resume,
        coverLetter,
        location,
        jd_source_path,
        job_category=job_category,
        position_type=position_type,
    )

    # Reconstruct the cover letter path (same logic as inside create_application_folder)
    parent_folder = os.path.dirname(resume)
    grandparent_folder = os.path.dirname(parent_folder)
    company_folder = os.path.join(grandparent_folder, company)
    position_folder = os.path.join(company_folder, position)
    cover_filename = f"Roger Xu_{company}_CoverLetter_Template.docx"
    cover_target = os.path.join(position_folder, cover_filename)

    # If not easy apply, run coverletter customizer
    if easy_apply == "false" and folder_created:
        print("🔄 Running coverletter customizer...")
        coverletter_script = "/Users/Roger/Documents/PersonalProject/Job Automation/coverletter_customizer.py"
        try:
            subprocess.run(f"python3 \"{coverletter_script}\"", shell=True, check=True)
            print("✅ Cover letter customized successfully")

            # Export cover letter as PDF (after customization)
            print("📄 Exporting cover letter to PDF...")
            try:
                # coverletter_customizer.py should produce non-template file
                customized_cover_path = cover_target.replace("_Template.docx", ".docx")
                if os.path.exists(customized_cover_path):
                    convert_docx_to_pdf(customized_cover_path)
                else:
                    convert_docx_to_pdf(cover_target)
            except Exception as e:
                print(f"⚠️ PDF export failed (non-critical): {e}")
                print("   Cover letter DOCX file is still available for manual PDF export.")
        except subprocess.CalledProcessError as e:
            print(f"❌ Failed to customize cover letter: {e}")
    elif folder_created:
        # Easy apply: still export cover letter as PDF (template-based)
        print("📄 Exporting cover letter to PDF (easy apply)...")
        try:
            convert_docx_to_pdf(cover_target)
        except Exception as e:
            print(f"⚠️ PDF export failed (non-critical): {e}")
            print("   Cover letter DOCX file is still available for manual PDF export.")
