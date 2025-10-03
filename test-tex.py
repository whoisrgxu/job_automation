import os
import shutil
import sys
from docx import Document
from datetime import datetime
from resume_customizer import customize_resume_with_placeholders
from openpyxl import load_workbook, Workbook
from pathlib import Path
from customize_resume_sections_tex import customize_resume_sections_tex


def already_applied(excel_path, sheet_name, company_name, position_name):
    """Check if a company/position already exists in the tracker."""
    if not os.path.exists(excel_path):
        return False  # no log file yet → definitely not applied

    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        return False  # sheet doesn’t exist yet

    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        if not row: 
            continue
        existing_company, existing_position, *_ = row
        if (existing_company and existing_company.strip().lower() == company_name.strip().lower() and
            existing_position and existing_position.strip().lower() == position_name.strip().lower()):
            return True
    return False

def log_application_to_excel(excel_path, sheet_name, company_name, position_name, applied_date):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(["Company", "Position", "Applied Date"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Company", "Position", "Applied Date"])

    # Append
    ws.append([company_name, position_name, applied_date])
    print(f"➡️ Logging row: {company_name}, {position_name}, {applied_date}")  # debug

    wb.save(excel_path)
    print(f"💾 Saved to {excel_path}")


def replace_placeholders_in_docx(input_path, output_path, replacements: dict):
    """Replace placeholders like {{COMPANY_NAME}} in a Word document."""
    doc = Document(input_path)
    for para in doc.paragraphs:
        for key, value in replacements.items():
            if key in para.text:
                for run in para.runs:
                    if key in run.text:
                        run.text = run.text.replace(key, value)
    doc.save(output_path)


def create_application_folder(
        company_name, 
        position_name, 
        position_type, 
        resume_template_folder_path, 
        coverLetter_path,
        jd_source_path=None):

    excel_log_path = "/Users/Roger/Documents/FullTime-Resume/Job Tracker.xlsx"

    if already_applied(excel_log_path, "Job Tracker", company_name, position_name):
        print(f"⚠️ Already applied to '{position_name}' @ '{company_name}'. Skipping...")
        return

    parent_folder = os.path.dirname(resume_template_folder_path)
    grandparent_folder = os.path.dirname(parent_folder)

    company_folder = os.path.join(grandparent_folder, company_name)
    os.makedirs(company_folder, exist_ok=True)
    position_folder = os.path.join(company_folder, position_name)
    os.makedirs(position_folder, exist_ok=True)

    # ---- copy the whole LaTeX template folder ----
    resume_folder_name = f"Roger_Xu_{company_name}_Resume_Folder"
    target_resume_folder = os.path.join(position_folder, resume_folder_name)
    shutil.copytree(resume_template_folder_path, target_resume_folder, dirs_exist_ok=True)

    cover_filename = f"Roger Xu_{company_name}_CoverLetter.docx"
    jd_filename = "job_description.txt"
    cover_target = os.path.join(position_folder, cover_filename)
    jd_target = os.path.join(position_folder, jd_filename)

    today_str = datetime.today().strftime("%B %d, %Y")
    replacements = {"{{COMPANY_NAME}}": company_name,
                    "{{POSITION_NAME}}": position_name,
                    "{{TODAY_DATE}}": today_str}

    # ---- JD load (same as before) ----
    job_description, additional_info = None, None
    if jd_source_path and os.path.exists(jd_source_path):
        shutil.copy(jd_source_path, jd_target)
        with open(jd_target, "r") as f:
            job_description = f.read()
        addl_source_path = os.path.join(os.path.dirname(jd_source_path), "additional_info.txt")
        if os.path.exists(addl_source_path):
            with open(addl_source_path, "r") as f:
                additional_info = f.read()
            print("ℹ️ Loaded additional_info.txt")

    # ---- choose section source files (same structure you used) ----
    position_type_folder_name = position_type.capitalize() + "_Sections"
    section_files = {
        "SUMMARY": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/summary.txt",
        "HOOPP_EXPERIENCE": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/hoopp_experience.txt",
        "PORTFOLIO_TRACKER": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/portfolio_tracker.txt",
        "SKILLS": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/skills.txt",
        "JOBPILOT": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/jobpilot.txt",
    }

    # ---- NEW: write LaTeX snippets into target_resume_folder/sections ----
    if job_description:
        customize_resume_sections_tex(
            template_folder=Path(target_resume_folder),
            section_files=section_files,
            job_description=job_description,
            additional_info=additional_info,
            compile_pdf=True,                           # set False if you’ll compile manually
            final_pdf_name=f"Roger_Xu_{company_name}_{position_name}.pdf"
        )

    # cover letter (unchanged)
    replace_placeholders_in_docx(coverLetter_path, cover_target, replacements)

    print(f"✅ Application folder: {position_folder}")
    print(f"📄 Files created:\n- {target_resume_folder}\n- {cover_target}\n- {jd_target}")

    # log + clear JD (unchanged)
    log_application_to_excel(excel_log_path, "Job Tracker", company_name, position_name, today_str)
    try:
        if jd_source_path and os.path.exists(jd_source_path):
            open(jd_source_path, "w").close()
        addl_source_path = os.path.join(os.path.dirname(jd_source_path), "additional_info.txt")
        if os.path.exists(addl_source_path):
            open(addl_source_path, "w").close()
    except Exception as e:
        print(f"⚠️ Failed to clear source job files: {e}")




if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python script.py <CompanyName> <PositionName> <Type>")
        sys.exit(1)

    # Resume templates
    resume_default = "/Users/Roger/Documents/FullTime-Resume/Rong Gang Xu_Resume_v3.docx"
    resume_frontend = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/Roger_resume_frontend_placeholder"
    resume_fullstack = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/Roger_resume_fullstack_placeholder"
    # resume_sharepoint = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/Roger Xu_SharePoint_Resume.docx"
    coverLetter = "/Users/Roger/Documents/FullTime-Resume/Roger Xu_coverletter.docx"

    jd_source_path = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/job_description.txt"

    # CLI args
    company = sys.argv[1].strip()
    position = sys.argv[2].strip()
    position_type = sys.argv[3]

    if position_type == "default":
        resume_folder_placeholder = resume_default
    elif position_type == "frontend":
        resume_folder_placeholder = resume_frontend
    elif position_type == "fullstack":
        resume_folder_placeholder = resume_fullstack
    # elif position_type == "sharepoint":
    #     resume = resume_sharepoint
    else:
        print("Invalid position type. Use 'default', 'frontend', 'fullstack', or 'sharepoint'.")
        sys.exit(1)

    create_application_folder(company, position, position_type, resume_folder_placeholder, coverLetter, jd_source_path)
