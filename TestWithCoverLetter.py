import os
import shutil
import sys
from docx import Document
from datetime import datetime
from resume_customizer import customize_resume_with_placeholders
from coverletter_customizer import customize_cover_letter
from openpyxl import load_workbook, Workbook


def already_applied(excel_path, sheet_name, company_name, position_name):
    """Check if a company/position already exists in the tracker."""
    if not os.path.exists(excel_path):
        return False

    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        return False

    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        existing_company, existing_position, *_ = row
        if (existing_company and existing_company.strip().lower() == company_name.strip().lower()
            and existing_position and existing_position.strip().lower() == position_name.strip().lower()):
            return True
    return False


def log_application_to_excel(excel_path, sheet_name, company_name, position_name, applied_date):
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(["Company", "Position", "Applied Date"])
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Company", "Position", "Applied Date"])

    ws.append([company_name, position_name, applied_date])
    print(f"➡️ Logging row: {company_name}, {position_name}, {applied_date}")

    wb.save(excel_path)
    print(f"💾 Saved to {excel_path}")


def replace_placeholders_in_docx(input_path, output_path, replacements: dict):
    """Replace placeholders like {{COMPANY_NAME}} in a Word document."""
    doc = Document(input_path)
    for para in doc.paragraphs:
        for key, value in replacements.items():
            if key in para.text:
                for run in para.runs:
                    if key in run.text:
                        run.text = run.text.replace(key, value)
    doc.save(output_path)


def create_application_folder(company_name, position_name, position_type,
                              resume_path, coverLetter_path, jd_source_path=None,
                              customize_cover=False):

    excel_log_path = "/Users/Roger/Documents/FullTime-Resume/Job Tracker.xlsx"

    if already_applied(excel_log_path, "Job Tracker", company_name, position_name):
        print(f"⚠️ WARNING: You have already applied to '{position_name}' at '{company_name}'. Skipping...")
        return

    parent_folder = os.path.dirname(resume_path)
    grandparent_folder = os.path.dirname(parent_folder)

    company_folder = os.path.join(grandparent_folder, company_name)
    os.makedirs(company_folder, exist_ok=True)
    position_folder = os.path.join(company_folder, position_name)
    os.makedirs(position_folder, exist_ok=True)

    resume_filename = f"Roger Xu_{company_name}_Resume.docx"
    cover_filename = f"Roger Xu_{company_name}_CoverLetter.docx"
    jd_filename = "job_description.txt"

    resume_target = os.path.join(position_folder, resume_filename)
    cover_target = os.path.join(position_folder, cover_filename)
    jd_target = os.path.join(position_folder, jd_filename)

    today_str = datetime.today().strftime("%B %d, %Y")

    replacements = {
        "{{COMPANY_NAME}}": company_name,
        "{{POSITION_NAME}}": position_name,
        "{{TODAY_DATE}}": today_str
    }

    shutil.copy(resume_path, resume_target)

    job_description = None
    additional_info = None

    if jd_source_path and os.path.exists(jd_source_path):
        shutil.copy(jd_source_path, jd_target)
        with open(jd_target, "r") as f:
            job_description = f.read()

        addl_source_path = os.path.join(os.path.dirname(jd_source_path), "additional_info.txt")
        if os.path.exists(addl_source_path):
            with open(addl_source_path, "r") as f:
                additional_info = f.read()
            print("ℹ️ Loaded additional_info.txt")

    position_type_folder_name = position_type.capitalize() + "_Sections"
    if job_description:
        section_files = {
            "SUMMARY": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/summary.txt",
            "HOOPP_EXPERIENCE": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/hoopp_experience.txt",
            "PORTFOLIO_TRACKER": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/portfolio_tracker.txt",
            "SKILLS": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/skills.txt",
            "JOBPILOT": f"/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/{position_type_folder_name}/jobpilot.txt",
        }

        customized_resume_path = os.path.join(
            position_folder,
            f"Roger Xu_{company_name}_Resume_Customized.docx"
        )
        customize_resume_with_placeholders(
            resume_path,
            section_files,
            job_description,
            customized_resume_path,
            additional_info
        )
        resume_target = customized_resume_path

    replace_placeholders_in_docx(coverLetter_path, cover_target, replacements)

    if customize_cover and job_description:
        customized_cover_path = os.path.join(
            position_folder, f"Roger Xu_{company_name}_CoverLetter_Customized.docx"
        )
        customize_cover_letter(cover_target, customized_cover_path, job_description, additional_info)
        cover_target = customized_cover_path
        print(f"✨ Customized cover letter created: {cover_target}")

    print(f"✅ Application folder created: {position_folder}")
    print(f"📄 Files created:\n- {resume_target}\n- {cover_target}\n- {jd_target}")

    log_application_to_excel(
        excel_log_path,
        sheet_name="Job Tracker",
        company_name=company_name,
        position_name=position_name,
        applied_date=today_str
    )
    print(f"📝 Logged application to {excel_log_path}")

    try:
        if jd_source_path and os.path.exists(jd_source_path):
            open(jd_source_path, "w").close()
            print("🧹 Cleared source job_description.txt content")

        addl_source_path = os.path.join(os.path.dirname(jd_source_path), "additional_info.txt")
        if os.path.exists(addl_source_path):
            open(addl_source_path, "w").close()
            print("🧹 Cleared source additional_info.txt content")
    except Exception as e:
        print(f"⚠️ Failed to clear source job description/additional_info: {e}")


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python script.py <CompanyName> <PositionName> <Type> [--customize_cover]")
        sys.exit(1)

    resume_default = "/Users/Roger/Documents/FullTime-Resume/Rong Gang Xu_Resume_v3.docx"
    resume_frontend = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/Roger Xu_Frontend_Resume_Placeholder.docx"
    resume_fullstack = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/Roger Xu_Fullstack_Resume_Placeholder.docx"
    resume_sharepoint = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/Roger Xu_SharePoint_Resume.docx"
    coverLetter = "/Users/Roger/Documents/FullTime-Resume/Roger Xu_coverletter.docx"

    jd_source_path = "/Users/Roger/Documents/FullTime-Resume/Resume Template - One Page/job_description.txt"

    company = sys.argv[1].strip()
    position = sys.argv[2].strip()
    position_type = sys.argv[3].lower()
    customize_cover = "--coverletter" in sys.argv

    if position_type == "default":
        resume = resume_default
    elif position_type == "frontend":
        resume = resume_frontend
    elif position_type == "fullstack":
        resume = resume_fullstack
    elif position_type == "sharepoint":
        resume = resume_sharepoint
    else:
        print("Invalid position type. Use 'default', 'frontend', 'fullstack', or 'sharepoint'.")
        sys.exit(1)

    create_application_folder(company, position, position_type, resume, coverLetter, jd_source_path, customize_cover)
