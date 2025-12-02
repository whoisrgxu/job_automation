"""
Utility for checking whether a job has already been applied to,
based on entries stored in the Job Tracker Excel workbook.
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook
from rapidfuzz import fuzz


DEFAULT_TRACKER_PATH = Path("/Users/Roger/Documents/FullTime-Resume/Job Tracker.xlsx")
DEFAULT_SHEET_NAME = "Job Tracker"
DEFAULT_LOOKBACK_DAYS = 90  # ~ 2 months
DEFAULT_SIMILARITY_THRESHOLD = 92


def _parse_tracker_date(value) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value
    if not value:
        return None
    if isinstance(value, str):
        cleaned = re.sub(r"(\d+)(st|nd|rd|th)", r"\1", value.strip())
        for fmt in ("%B %d, %Y", "%b %d, %Y"):
            try:
                return datetime.strptime(cleaned, fmt)
            except ValueError:
                continue
    return None


def _build_header_map(ws) -> Dict[str, int]:
    headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {str(header).strip().lower(): idx for idx, header in enumerate(headers) if header}


def _get_cell(row: tuple, index: Optional[int]):
    if index is None:
        return None
    return row[index] if index < len(row) else None


@dataclass
class TrackerEntry:
    company: str
    position: str
    description: str
    applied_date: Optional[datetime]


class AppliedTracker:
    def __init__(
        self,
        excel_path: Path = DEFAULT_TRACKER_PATH,
        sheet_name: str = DEFAULT_SHEET_NAME,
        lookback_days: int = DEFAULT_LOOKBACK_DAYS,
        similarity_threshold: int = DEFAULT_SIMILARITY_THRESHOLD,
    ):
        self.excel_path = Path(excel_path)
        self.sheet_name = sheet_name
        self.lookback_days = lookback_days
        self.similarity_threshold = similarity_threshold
        self.entries: List[TrackerEntry] = []
        self._load_entries()

    def _load_entries(self) -> None:
        if not self.excel_path.exists():
            print(f"⚠️ Tracker not found at {self.excel_path}. Skipping duplicate filtering.")
            self.entries = []
            return

        try:
            wb = load_workbook(self.excel_path, data_only=True)
        except Exception as exc:
            print(f"⚠️ Failed to open tracker {self.excel_path}: {exc}")
            self.entries = []
            return

        if self.sheet_name not in wb.sheetnames:
            print(f"⚠️ Sheet '{self.sheet_name}' not found in tracker. Skipping duplicate filtering.")
            self.entries = []
            return

        ws = wb[self.sheet_name]
        header_map = _build_header_map(ws)

        company_idx = header_map.get("company", 0)
        position_idx = header_map.get("position", 1)
        date_idx = header_map.get("applied date", 2)
        description_idx = header_map.get("job description")

        cutoff = datetime.now() - timedelta(days=self.lookback_days)
        entries: List[TrackerEntry] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            company_raw = _get_cell(row, company_idx)
            position_raw = _get_cell(row, position_idx)
            if not company_raw or not position_raw:
                continue

            applied_date = _parse_tracker_date(_get_cell(row, date_idx))
            if applied_date and applied_date < cutoff:
                continue

            description_val = _get_cell(row, description_idx) or ""

            entries.append(
                TrackerEntry(
                    company=str(company_raw).strip().lower(),
                    position=str(position_raw).strip().lower(),
                    description=str(description_val),
                    applied_date=applied_date,
                )
            )

        self.entries = entries

    def is_applied(
        self,
        company_name: str,
        position_name: str,
        job_description: Optional[str] = None,
    ) -> bool:
        if not self.entries:
            return False

        company_key = (company_name or "").strip().lower()
        position_key = (position_name or "").strip().lower()
        snippet_current = (job_description or "")[:800]

        # Check if the company and position match with job description match check
        for entry in self.entries:
            if entry.company == company_key and entry.position == position_key:
                if snippet_current and entry.description:
                    similarity = fuzz.token_set_ratio(snippet_current, entry.description[:800])
                    if similarity >= self.similarity_threshold:
                        return True
                else:
                    return True
        return False

